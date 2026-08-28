from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from infrastructure.config import (
    OUTPUT_SHEET_NAME,
    OUTPUT_COLUMNS,
    OUTPUT_START_ROW,
    COLUMN_WIDTHS,
    YELLOW_FILL_COLOR,
    BLUE_FILL_COLOR,
    YELLOW_COLUMNS,
    HEADER_ROW_OUTPUT,
)
from domain.model import EmployeeMetric
from datetime import date
from openpyxl.styles import PatternFill
from infrastructure.config import (
    COLORS,
    PASSIVE_COLUMN,
    HOURS_DAY_COLUMN,
    PRODUCTIVE_ACTIVE_COLUMN,
    TOTAL_PRODUCTIVE_COLUMN,
    WEEK_ROW_OUTPUT,
)


def create_workbook():

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = OUTPUT_SHEET_NAME

    return workbook, worksheet


def write_week_range(
    worksheet, week_start: date, week_end: date, start_row: int
) -> None:

    week_text = (
        f"Week: "
        f"{week_start.strftime('%m/%d/%Y')} - "
        f"{week_end.strftime('%m/%d/%Y')}"
    )

    week_cell = worksheet.cell(row=start_row, column=1, value=week_text)

    week_font = Font(bold=True, color="FFFFFF", size=16)

    week_fill = PatternFill(fill_type="solid", fgColor="FF000000")

    week_cell.font = week_font
    week_cell.fill = week_fill
    week_cell.alignment = Alignment(horizontal="center", vertical="center")

    last_column_letter = get_column_letter(len(OUTPUT_COLUMNS))

    worksheet.merge_cells(f"A{start_row}:{last_column_letter}{start_row}")


def write_headers(worksheet, start_row: int) -> None:
    for column, header in enumerate(OUTPUT_COLUMNS, start=1):
        worksheet.cell(row=start_row + 1, column=column, value=header)


def write_employees(worksheet, employees: list[EmployeeMetric], start_row: int) -> None:

    for row_number, employee in enumerate(employees, start=start_row + 2):
        values = [
            employee.name,
            employee.department,
            employee.productive_active_hours,
            employee.productive_passive_hours,
            employee.total_hours,
            employee.active_days,
            employee.goal,
            employee.hours_per_day,
            employee.comments,
        ]

        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_number, column=column, value=value)

            if column == HOURS_DAY_COLUMN:
                cell.number_format = "0.0"

            if column == PRODUCTIVE_ACTIVE_COLUMN:
                cell.number_format = "0.0"

            if column == PASSIVE_COLUMN:
                cell.number_format = "0.0"

            if column == TOTAL_PRODUCTIVE_COLUMN:
                cell.number_format = "0.0"


def apply_colors(worksheet, employees: list[EmployeeMetric], start_row: int) -> None:

    green_fill = PatternFill(fill_type="solid", fgColor=COLORS["green"])

    red_fill = PatternFill(fill_type="solid", fgColor=COLORS["red"])

    yellow_fill = PatternFill(fill_type="solid", fgColor=COLORS["yellow"])

    for row_number, employee in enumerate(employees, start=start_row + 2):

        if employee.color_hours_day == "green":
            worksheet.cell(row=row_number, column=HOURS_DAY_COLUMN).fill = green_fill

        elif employee.color_hours_day == "yellow":
            worksheet.cell(row=row_number, column=HOURS_DAY_COLUMN).fill = yellow_fill

        elif employee.color_hours_day == "red":
            worksheet.cell(row=row_number, column=HOURS_DAY_COLUMN).fill = red_fill

        elif employee.color_hours_day == "none":
            pass

        if employee.color_passive == "green":
            worksheet.cell(row=row_number, column=PASSIVE_COLUMN).fill = green_fill

        elif employee.color_passive == "red":
            worksheet.cell(row=row_number, column=PASSIVE_COLUMN).fill = red_fill

        elif employee.color_passive == "none":
            pass

        if employee.color_active_hrs == "green":
            worksheet.cell(row=row_number, column=PRODUCTIVE_ACTIVE_COLUMN).fill = green_fill

        elif employee.color_active_hrs == "red":
            worksheet.cell(row=row_number, column=PRODUCTIVE_ACTIVE_COLUMN).fill = red_fill
        
        elif employee.color_active_hrs == "none":
            pass

        if employee.color_total_hours == "green":
            worksheet.cell(row=row_number, column=TOTAL_PRODUCTIVE_COLUMN).fill = green_fill

        elif employee.color_total_hours == "yellow":
            worksheet.cell(row=row_number, column=TOTAL_PRODUCTIVE_COLUMN).fill = yellow_fill

        elif employee.color_total_hours == "red":
            worksheet.cell(row=row_number, column=TOTAL_PRODUCTIVE_COLUMN).fill = red_fill
        
        elif employee.color_total_hours == "none":
            pass

def set_column_width(worksheet) -> None:
    for column, width in COLUMN_WIDTHS.items():
        letter = get_column_letter(column)
        worksheet.column_dimensions[letter].width = width


def apply_column_colors(worksheet, start_row: int) -> None:

    yellow_fill = PatternFill(fill_type="solid", fgColor=YELLOW_FILL_COLOR)

    blue_fill = PatternFill(fill_type="solid", fgColor=BLUE_FILL_COLOR)

    for column in range(1, len(OUTPUT_COLUMNS) + 1):

        if column in YELLOW_COLUMNS:
            fill = yellow_fill
        else:
            fill = blue_fill

        worksheet.cell(row=start_row + 1, column=column).fill = fill


def apply_header_font(worksheet, start_row: int) -> None:

    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    worksheet.row_dimensions[start_row + 1].height = 25

    for column in range(1, len(OUTPUT_COLUMNS) + 1):
        cell = worksheet.cell(row=start_row + 1, column=column)

        cell.font = header_font
        cell.alignment = header_alignment


def apply_table_borders(
    worksheet, employees: list[EmployeeMetric], start_row: int
) -> None:

    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    last_row = start_row + len(employees) + 1

    for row in range(start_row, last_row + 1):
        for column in range(1, len(OUTPUT_COLUMNS) + 1):
            worksheet.cell(row=row, column=column).border = thin_border


def get_next_block_start_row(worksheet) -> int:
    if worksheet.max_row <= 1:
        return WEEK_ROW_OUTPUT
    return worksheet.max_row + 4


def week_already_exists(worksheet, week_start: date, week_end: date) -> bool:

    expected_text = (
        f"Week: "
        f"{week_start.strftime('%m/%d/%Y')} - "
        f"{week_end.strftime('%m/%d/%Y')}"
    )

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == expected_text:
                return True

    return False


