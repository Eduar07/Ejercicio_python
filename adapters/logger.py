import io
from openpyxl import Workbook
import openpyxl

from domain.model import LogEntry
from adapters.storage import get_log_path
from adapters.sharepoint_client import (
    GraphContext,
    get_file_by_path,
    upload_file_graph,
)
from infrastructure.config import LOG_COLUMNS
from infrastructure.graph_config import BASE_FOLDER


def write_log_entry(
    log_entry: LogEntry,
) -> None:

    log_path = get_log_path(
        log_entry.execution_date.year
    )

    if log_path.exists():
        workbook = openpyxl.load_workbook(log_path)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active

        for column, header in enumerate(
            LOG_COLUMNS,
            start=1,
        ):
            worksheet.cell(
                row=1,
                column=column,
                value=header,
            )

    next_row = worksheet.max_row + 1

    values = [
        log_entry.source_file,
        log_entry.execution_date.strftime(
            "%Y-%m-%d %H:%M:%S"
        ),
        log_entry.status,
        log_entry.error_message,
        log_entry.error_code,
    ]

    for column, value in enumerate(
        values,
        start=1,
    ):
        worksheet.cell(
            row=next_row,
            column=column,
            value=value,
        )

    workbook.save(log_path)


def write_log_entry_remote(
    context: GraphContext,
    log_entry: LogEntry,
) -> None:

    folder = f"{BASE_FOLDER}/Parametric Files"

    file_name = (
        f"AT_Process_Log_"
        f"{log_entry.execution_date.year}.xlsx"
    )

    path = f"{folder}/{file_name}"

    existing_bytes = get_file_by_path(
        context,
        path,
    )

    if existing_bytes is not None:
        workbook = openpyxl.load_workbook(
            existing_bytes
        )
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active

        for column, header in enumerate(
            LOG_COLUMNS,
            start=1,
        ):
            worksheet.cell(
                row=1,
                column=column,
                value=header,
            )

    next_row = worksheet.max_row + 1

    values = [
        log_entry.source_file,
        log_entry.execution_date.strftime(
            "%Y-%m-%d %H:%M:%S"
        ),
        log_entry.status,
        log_entry.error_message,
        log_entry.error_code,
    ]

    for column, value in enumerate(
        values,
        start=1,
    ):
        worksheet.cell(
            row=next_row,
            column=column,
            value=value,
        )

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    upload_file_graph(
        context,
        folder,
        file_name,
        buffer.getvalue(),
    )