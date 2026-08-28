"""
Adapter for reading the master employees file (AT_Employees.xlsx).
"""
import io
from domain.model import MasterEmployee
from pathlib import Path
import openpyxl
from infrastructure.config import SHEET_MASTER, START_ROW_MASTER, HEADER_ROW_MASTER
from domain.errors import ATError
from infrastructure.config import MASTER_COLUMNS


def validate_master_headers(headers_master: list[str]) -> None:
    if headers_master != MASTER_COLUMNS:
        raise ATError(
            "ERR006", "Excel file columns do not match the expected structure"
        )


def get_master_headers(worksheetmaster) -> list[str]:
    headers_master = []
    for row in worksheetmaster.iter_rows(
        min_row=HEADER_ROW_MASTER, max_row=HEADER_ROW_MASTER
    ):
        for cell in row:
            headers_master.append(cell.value)
    return headers_master


def create_column_map(headers: list[str]) -> dict[str, int]:
    column_map = {}
    for index, header in enumerate(headers):
        column_map[header] = index

    return column_map


def parse_master_employees(
    worksheetmaster, column_map_master: dict[str, int]
) -> list[MasterEmployee]:

    master = []

    for row in worksheetmaster.iter_rows(min_row=START_ROW_MASTER):

        employee_id = row[column_map_master["EmployeeID"]].value

        if employee_id is None:
            continue

        master_employee = MasterEmployee(
            employee_id=employee_id,
            name=row[column_map_master["Name"]].value,
            email=row[column_map_master["Email"]].value,
            status=row[column_map_master["Status"]].value,
        )

        master.append(master_employee)

    return master


def read_excel(file_path: Path | io.BytesIO, source_name: str | None = None) -> list[MasterEmployee]:
    try:
        workbook = openpyxl.load_workbook(file_path)
    except Exception:
        raise ATError("ERR004", "Could not open the file")

    try:
        worksheetmaster = workbook[SHEET_MASTER]
    except KeyError:
        raise ATError("ERR005", f"Sheet {SHEET_MASTER} does not exist")

    headers_master = get_master_headers(worksheetmaster)
    validate_master_headers(headers_master)
    column_map_master = create_column_map(headers_master)
    master = parse_master_employees(worksheetmaster, column_map_master)

    return master
