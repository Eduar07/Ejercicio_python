"""
Adapter implementing ExecutionLogPort against SharePoint (via Graph) —
writes execution outcomes to a yearly Excel log file.
"""
import io
import openpyxl
from openpyxl import Workbook

from domain.model import LogEntry

from application.ports.execution_log_port import ExecutionLogPort

from infrastructure.config import LOG_COLUMNS
from infrastructure.graph_config import BASE_FOLDER

from adapters.sharepoint_client import GraphContext, get_file_by_path, upload_file_graph


class SharePointExecutionLogAdapter(ExecutionLogPort):
    def __init__(self, context: GraphContext):
        self._context = context

    def write(self, log_entry: LogEntry) -> None:
        folder = f"{BASE_FOLDER}/Parametric Files"
        file_name = f"AT_Process_Log_{log_entry.execution_date.year}.xlsx"
        path = f"{folder}/{file_name}"

        existing_bytes = get_file_by_path(self._context, path)

        if existing_bytes is not None:
            workbook = openpyxl.load_workbook(existing_bytes)
            worksheet = workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active

            for column, header in enumerate(LOG_COLUMNS, start=1):
                worksheet.cell(row=1, column=column, value=header)

        next_row = worksheet.max_row + 1

        values = [
            log_entry.source_file,
            log_entry.execution_date.strftime("%Y-%m-%d %H:%M:%S"),
            log_entry.status,
            log_entry.error_message,
            log_entry.error_code,
        ]

        for column, value in enumerate(values, start=1):
            worksheet.cell(row=next_row, column=column, value=value)

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        upload_file_graph(self._context, folder, file_name, buffer.getvalue())
