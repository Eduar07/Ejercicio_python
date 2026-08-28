"""
Adapter for reading the two raw ActivTrak exports directly:
- Productivity by User (hours: ProdActive secs, ProdPassive secs)
- User_Details (Active Days)
Produces the same EmployeeMetric list that business_rules.py and
excel_write.py already consume, so nothing downstream changes.
"""
import io

import openpyxl

from domain.model import EmployeeMetric
from domain.errors import ATError
from infrastructure.config import (
    RAW_HEADER_ROW,
    RAW_START_ROW,
    PROD_BY_USER_COLUMNS_NEEDED,
    USER_DETAILS_COLUMNS_NEEDED,
)


def get_headers_raw(worksheet) -> list[str]:
    headers = []
    for row in worksheet.iter_rows(min_row=RAW_HEADER_ROW, max_row=RAW_HEADER_ROW):
        for cell in row:
            headers.append(cell.value)
    return headers


def create_column_map_raw(headers: list[str]) -> dict[str, int]:
    column_map = {}
    for index, header in enumerate(headers):
        column_map[header] = index
    return column_map


def validate_raw_headers(headers: list[str], required_columns: list[str], file_label: str) -> None:
    missing = [col for col in required_columns if col not in headers]
    if missing:
        raise ATError(
            "ERR006",
            f"{file_label} is missing expected columns: {missing}"
        )


def read_prod_by_user_hours(file_bytes: io.BytesIO) -> dict[str, tuple[float, float]]:
    """
    Reads Productivity by User raw file.
    Returns {employee_name: (productive_active_hours, productive_passive_hours)}
    Converts secs -> decimal hours (secs / 3600).
    """
    try:
        workbook = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        raise ATError("ERR004", "Could not open Productivity by User file")

    worksheet = workbook.active

    headers = get_headers_raw(worksheet)
    validate_raw_headers(headers, PROD_BY_USER_COLUMNS_NEEDED, "Productivity by User")
    column_map = create_column_map_raw(headers)

    hours_by_name = {}

    for row in worksheet.iter_rows(min_row=RAW_START_ROW):
        name = row[column_map["User"]].value

        if name is None:
            continue

        active_secs = row[column_map["ProdActive (secs)"]].value or 0
        passive_secs = row[column_map["ProdPassive (secs)"]].value or 0

        active_hours = active_secs / 3600
        passive_hours = passive_secs / 3600

        hours_by_name[name] = (active_hours, passive_hours)

    return hours_by_name


def read_user_details_active_days(file_bytes: io.BytesIO) -> dict[str, int]:
    """
    Reads User_Details raw file.
    Returns {employee_name: active_days}
    """
    try:
        workbook = openpyxl.load_workbook(file_bytes, data_only=True)
    except Exception:
        raise ATError("ERR004", "Could not open User_Details file")

    worksheet = workbook.active

    headers = get_headers_raw(worksheet)
    validate_raw_headers(headers, USER_DETAILS_COLUMNS_NEEDED, "User_Details")
    column_map = create_column_map_raw(headers)

    active_days_by_name = {}

    for row in worksheet.iter_rows(min_row=RAW_START_ROW):
        name = row[column_map["User"]].value

        if name is None:
            continue

        active_days = row[column_map["Active Days"]].value or 0
        active_days_by_name[name] = active_days

    return active_days_by_name


def build_employee_metrics_from_raw(
    prod_by_user_bytes: io.BytesIO,
    user_details_bytes: io.BytesIO,
    source_file: str,
) -> list[EmployeeMetric]:
    """
    Combines both raw files by employee name into a list of EmployeeMetric,
    ready to be handed to business_rules.apply_business_rules().
    Department, GOAL, hours_per_day, and colors are filled in later —
    same as the existing pipeline already does.
    """
    hours_by_name = read_prod_by_user_hours(prod_by_user_bytes)
    active_days_by_name = read_user_details_active_days(user_details_bytes)

    employees = []

    all_names = set(hours_by_name.keys()) | set(active_days_by_name.keys())

    for name in all_names:
        active_hours, passive_hours = hours_by_name.get(name, (0, 0))
        active_days = active_days_by_name.get(name, 0)

        total_hours = active_hours + passive_hours
        hours_per_day = total_hours / active_days if active_days else 0

        employee = EmployeeMetric(
            name=name,
            department="",  # filled in by match_employees() later
            productive_active_hours=active_hours,
            productive_passive_hours=passive_hours,
            total_hours=total_hours,
            active_days=active_days,
            goal=0,  # calculated by apply_business_rules()
            hours_per_day=hours_per_day,
            comments="",
            source_file=source_file,
        )

        employees.append(employee)

    return employees
